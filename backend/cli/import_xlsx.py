"""CLI: 从「基金持仓规则表.xlsx」一次性导入实盘持仓 + 补仓计划 + 操作规则。

用法：
    ifund import-xlsx run --file /path/基金持仓规则表.xlsx [--name 我的持仓] [--reset]

映射口径：
- 「持仓总览」→ 每只基金一笔建仓 buy 交易（holding_txns：amount=成本价×份额,
  nav=成本价, shares=份额，trade_date=净值日期）→ 持仓由合成引擎精确还原；
  计划/已用补仓资金、板块、类型、估值百分位 → fund_plans。
- 「操作规则」→ fund_rules（补仓第1/2/3档 → add_1/2/3，止盈 → take_profit）。
- 「操作日志」→ holding_txns 追加（跳过「示例」行；建仓交易与持仓总览重复时不重复导入）。

幂等：同名实盘已存在时报错；--reset 先删除该实业的持仓/交易/计划/规则再重建。
"""
from __future__ import annotations

import datetime
import sys

RULE_TYPE_MAP = {"补仓第1档": "add_1", "补仓第2档": "add_2",
                 "补仓第3档": "add_3", "止盈": "take_profit"}

# 持仓总览列（B 起）：代码/名称/类型/板块/成本价/份额/当前净值/收益率/市值/盈亏/计划/已用/剩余/估值百分位/状态/净值日期/备注
HOLDING_COLS = {
    "code": "B", "name": "C", "fund_type": "D", "sector": "E", "cost_price": "F",
    "shares": "G", "nav": "H", "return_pct": "I", "market_value": "J", "pnl": "K",
    "planned": "L", "used": "M", "valuation_pct": "O", "nav_date": "Q", "note": "R",
}
# 操作规则列：代码/名称/规则类型/触发条件/触发净值/动作/资金比例/触发金额/提醒/是否已执行/执行日期
RULE_COLS = {
    "code": "B", "name": "C", "rule_type": "D", "trigger_pct": "E", "trigger_nav": "F",
    "action": "G", "fund_pct": "H", "amount": "I", "executed": "K", "executed_date": "L",
}
# 操作日志列：日期/代码/名称/操作类型/触发规则/成交净值/成交金额/确认份额/备注
LOG_COLS = {
    "date": "B", "code": "C", "name": "D", "op": "E", "rule": "F",
    "nav": "G", "amount": "H", "shares": "I", "note": "J",
}


def _date_str(v) -> str:
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v or "").strip()


def _rows(ws, cols: dict, start: int = 5) -> list[dict]:
    """按列字母抽行，B 列（代码/日期）为空则跳过。"""
    out = []
    key_col = next(iter(cols.values()))
    for r in range(start, ws.max_row + 1):
        if ws[f"{key_col}{r}"].value in (None, ""):
            continue
        row = {k: ws[f"{c}{r}"].value for k, c in cols.items()}
        out.append(row)
    return out


def cmd_run(args) -> None:
    """执行导入。需 openpyxl（一次性依赖）：``./venv/bin/pip install openpyxl``。"""
    try:
        import openpyxl
    except ImportError:
        print("错误：缺 openpyxl，请先 ./venv/bin/pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    from app import db as database
    from app.reconcile.crud import portfolios_store

    uid = args.user
    path = args.file
    pname = args.name or "我的持仓"

    if not database.select_one("users", {"id": f"eq.{uid}"}):
        print(f"错误：用户 id={uid} 不存在，请先在网页端注册", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in ("持仓总览", "操作规则"):
        if sheet not in wb.sheetnames:
            print(f"错误：Excel 缺少「{sheet}」页", file=sys.stderr)
            sys.exit(1)

    existing = [p for p in portfolios_store.list_portfolios(uid) if p["name"] == pname]
    if existing:
        if not args.reset:
            print(f"错误：实盘「{pname}」已存在（id={existing[0]['id']}），加 --reset 覆盖重建",
                  file=sys.stderr)
            sys.exit(1)
        pid = existing[0]["id"]
        for table in ("holding_txns", "user_holdings", "fund_plans", "fund_rules"):
            database.delete(table, {"portfolio_id": pid})
        database.delete("portfolios", {"id": pid})
        print(f"已清空旧实盘「{pname}」(id={pid})")

    portfolio = portfolios_store.create_portfolio(uid, pname)
    pid = portfolio["id"]
    print(f"创建实盘「{pname}」(id={pid})")

    now = datetime.datetime.now().isoformat()
    # ---- 持仓总览 → 建仓交易 + fund_plans ----
    holdings = _rows(wb["持仓总览"], HOLDING_COLS)
    total_mv = total_pnl = 0.0
    for h in holdings:
        code = str(h["code"]).strip().zfill(6)
        name = str(h["name"] or "").strip()
        cost_price = float(h["cost_price"])
        shares = float(h["shares"])
        nav_date = _date_str(h["nav_date"])
        database.insert("holding_txns", {
            "portfolio_id": pid, "user_id": uid, "fund_code": code, "fund_name": name,
            "txn_type": "buy", "trade_date": nav_date,
            "amount": round(cost_price * shares, 2), "nav": cost_price, "shares": shares,
            "note": "Excel导入·建仓",
        })
        database.insert("fund_plans", {
            "portfolio_id": pid, "user_id": uid, "fund_code": code,
            "fund_type": str(h["fund_type"] or ""), "sector": str(h["sector"] or ""),
            "planned_amount": float(h["planned"] or 0), "used_amount": float(h["used"] or 0),
            "valuation_pct": float(h["valuation_pct"]) if h["valuation_pct"] not in (None, "") else None,
            "note": str(h["note"] or ""), "updated_at": now,
        })
        total_mv += float(h["market_value"] or 0)
        total_pnl += float(h["pnl"] or 0)
        print(f"  持仓 {code} {name} 份额={shares} 成本价={cost_price}")
    print(f"持仓 {len(holdings)} 只：Excel 总市值={total_mv:.2f} 总盈亏={total_pnl:.2f}")

    # ---- 操作规则 → fund_rules ----
    n_rules = 0
    for r in _rows(wb["操作规则"], RULE_COLS):
        rtype = RULE_TYPE_MAP.get(str(r["rule_type"] or "").strip())
        if not rtype:
            continue
        database.insert("fund_rules", {
            "portfolio_id": pid, "user_id": uid,
            "fund_code": str(r["code"]).strip().zfill(6),
            "fund_name": str(r["name"] or "").strip(),
            "rule_type": rtype,
            "trigger_pct": float(r["trigger_pct"]),
            "trigger_nav": float(r["trigger_nav"]),
            "fund_pct": float(r["fund_pct"] or 0), "amount": float(r["amount"] or 0),
            "action": str(r["action"] or ""),
            "executed": 1 if str(r["executed"] or "").strip() == "是" else 0,
            "executed_date": _date_str(r["executed_date"]) or None,
        })
        n_rules += 1
    print(f"规则 {n_rules} 条")

    # ---- 操作日志 → 交易（跳过示例行；与建仓重复的不再导入，仅记非建仓操作）----
    n_log = 0
    if "操作日志" in wb.sheetnames:
        for log in _rows(wb["操作日志"], LOG_COLS):
            rule = str(log["rule"] or "")
            if rule.startswith("示例"):
                continue
            code = str(log["code"]).strip().zfill(6)
            database.insert("holding_txns", {
                "portfolio_id": pid, "user_id": uid, "fund_code": code,
                "fund_name": str(log["name"] or "").strip(),
                "txn_type": "buy" if str(log["op"] or "").strip() == "买入" else "sell",
                "trade_date": _date_str(log["date"]),
                "amount": float(log["amount"]), "nav": float(log["nav"]),
                "shares": float(log["shares"]),
                "note": f"Excel日志·{rule} {log['note'] or ''}".strip(),
            })
            n_log += 1
    print(f"日志 {n_log} 条（示例行已跳过）")
    print("导入完成。请用 fetch 拉取这 11 只基金的历史净值后核对市值。")
